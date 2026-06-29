"""Export docs/Pharmacy_list.xlsx to assets/data/pharmacies.json."""
import json
import re
from datetime import date
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "docs" / "Pharmacy_list.xlsx"
OUT = ROOT / "assets" / "data" / "pharmacies.json"


def main():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    pharmacies = []

    for row in range(4, ws.max_row + 1):
        row_num = ws.cell(row, 2).value
        name = ws.cell(row, 3).value
        if not name:
            continue
        stock = ws.cell(row, 8).value
        pharmacies.append(
            {
                "id": int(row_num) if row_num else len(pharmacies) + 1,
                "name": str(name).strip(),
                "city": str(ws.cell(row, 4).value or "").strip(),
                "phone": re.sub(r"\s+", "", str(ws.cell(row, 5).value or "").strip()),
                "region": str(ws.cell(row, 6).value or "").strip(),
                "address": str(ws.cell(row, 7).value or "").strip(),
                "stock": int(stock) if stock is not None else None,
            }
        )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    payload = {"updated_at": date.today().isoformat(), "pharmacies": pharmacies}
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {len(pharmacies)} pharmacies to {OUT}")


if __name__ == "__main__":
    main()
